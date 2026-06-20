"""
ADR Causality Assessment System - Flask Backend
Mobile PWA for Adverse Drug Reaction reporting and causality assessment.
Supports OCR scanning of IPC ADR forms via Google Cloud Vision + Gemini.
"""
import os
import json
import base64
import tempfile
import traceback
from datetime import datetime
from dotenv import load_dotenv

# Load environment variables from .env
load_dotenv()
print("DEBUG: Environment variables loaded. GEMINI_API_KEY present:", bool(os.environ.get('GEMINI_API_KEY')))

from flask import Flask, render_template, request, jsonify, redirect, url_for, Response, send_file
from flask_cors import CORS

# Google Cloud credentials from environment variable (for Render deployment)
google_creds_json = os.environ.get('GOOGLE_APPLICATION_CREDENTIALS_JSON')
if google_creds_json:
    creds_dict = json.loads(google_creds_json)
    tmp = tempfile.NamedTemporaryFile(mode='w', suffix='.json', delete=False)
    json.dump(creds_dict, tmp)
    tmp.close()
    os.environ['GOOGLE_APPLICATION_CREDENTIALS'] = tmp.name

from config import Config
from database import db, init_db
from database.models import ADRReport, SuspectedMedication, ConcomitantMedication, ModelTrainingLog
from ml.naranjo import calculate_naranjo_score, NARANJO_QUESTIONS, extract_naranjo_features_from_report
from ml.who_umc import assess_who_umc, assess_from_api_payload, WHO_UMC_CATEGORIES, run_assessment

# Configure Gemini (new google.genai SDK)
try:
    from google import genai
    from google.genai import types
    gemini_client = genai.Client(api_key=os.environ.get('GEMINI_API_KEY', ''))
    GEMINI_AVAILABLE = True
except Exception:
    GEMINI_AVAILABLE = False
    gemini_client = None

# Configure Google Cloud Vision (Disabled in favor of Gemini Multimodal OCR)
VISION_AVAILABLE = False


app = Flask(__name__)
app.config.from_object(Config)
CORS(app)

# Initialize database
init_db(app)

# Ensure directories exist
os.makedirs(app.config.get('UPLOAD_FOLDER', 'uploads'), exist_ok=True)


# ==================== PAGE ROUTES ====================

@app.route('/')
def index():
    """Dashboard / Home page."""
    with app.app_context():
        total_reports = ADRReport.query.count()
        serious_reports = ADRReport.query.filter_by(is_serious=True).count()
        recent_reports = ADRReport.query.order_by(ADRReport.created_at.desc()).limit(5).all()

    return render_template('index.html',
                           total_reports=total_reports,
                           serious_reports=serious_reports,
                           recent_reports=recent_reports)


@app.route('/form')
def adr_form():
    """ADR Reporting Form page."""
    return render_template('form.html')


@app.route('/reports')
def reports_list():
    """View all ADR reports."""
    page = request.args.get('page', 1, type=int)
    per_page = 20
    reports = ADRReport.query.order_by(ADRReport.created_at.desc()).paginate(
        page=page, per_page=per_page, error_out=False
    )
    return render_template('reports.html', reports=reports)


@app.route('/report/<int:report_id>')
def view_report(report_id):
    """View a single ADR report with causality assessment."""
    report = ADRReport.query.get_or_404(report_id)
    return render_template('report_detail.html', report=report)


@app.route('/analytics')
def analytics():
    """Analytics dashboard."""
    return render_template('analytics.html')


@app.route('/naranjo')
def naranjo_page():
    """Naranjo scale calculator page."""
    return render_template('naranjo.html', questions=NARANJO_QUESTIONS)


@app.route('/capture')
def capture():
    """Camera capture page for scanning ADR forms."""
    return render_template('capture.html')


# ==================== API ROUTES ====================

@app.route('/api/submit-report', methods=['POST'])
def submit_report():
    """Submit a new ADR report."""
    try:
        data = request.get_json()

        # Create main report
        report = ADRReport(
            case_type=data.get('case_type'),
            reg_no=data.get('reg_no'),
            amc_report_no=data.get('amc_report_no'),
            worldwide_unique_no=data.get('worldwide_unique_no'),
            patient_initials=data.get('patient_initials'),
            patient_age=data.get('patient_age'),
            patient_dob=data.get('patient_dob'),
            gender=data.get('gender'),
            weight_kg=float(data['weight_kg']) if data.get('weight_kg') else None,
            reaction_start_date=data.get('reaction_start_date'),
            reaction_stop_date=data.get('reaction_stop_date'),
            reaction_description=data.get('reaction_description'),
            relevant_investigations=data.get('relevant_investigations'),
            medical_history=data.get('medical_history'),
            is_serious=data.get('is_serious', False),
            seriousness_death=data.get('seriousness_death', False),
            seriousness_death_date=data.get('seriousness_death_date'),
            seriousness_life_threatening=data.get('seriousness_life_threatening', False),
            seriousness_hospitalization=data.get('seriousness_hospitalization', False),
            seriousness_congenital_anomaly=data.get('seriousness_congenital_anomaly', False),
            seriousness_disability=data.get('seriousness_disability', False),
            seriousness_other=data.get('seriousness_other', False),
            outcome=data.get('outcome'),
            additional_info=data.get('additional_info'),
            reporter_name=data.get('reporter_name'),
            reporter_address=data.get('reporter_address'),
            reporter_pin=data.get('reporter_pin'),
            reporter_email=data.get('reporter_email'),
            reporter_contact=data.get('reporter_contact'),
            reporter_occupation=data.get('reporter_occupation'),
            report_date=data.get('report_date'),
        )

        # Add suspected medications
        for med_data in data.get('medications', []):
            med = SuspectedMedication(
                serial_no=med_data.get('serial_no'),
                drug_name=med_data.get('drug_name'),
                manufacturer=med_data.get('manufacturer'),
                batch_no=med_data.get('batch_no'),
                expiry_date=med_data.get('expiry_date'),
                dose=med_data.get('dose'),
                route=med_data.get('route'),
                frequency=med_data.get('frequency'),
                therapy_start_date=med_data.get('therapy_start_date'),
                therapy_stop_date=med_data.get('therapy_stop_date'),
                indication=med_data.get('indication'),
                action_taken=med_data.get('action_taken'),
                reintroduction_result=med_data.get('reintroduction_result'),
                reintroduction_dose=med_data.get('reintroduction_dose'),
                causality_assessment=med_data.get('causality_assessment'),
            )
            report.medications.append(med)

        # Add concomitant medications
        for con_data in data.get('concomitant_medications', []):
            con = ConcomitantMedication(
                serial_no=con_data.get('serial_no'),
                drug_name=con_data.get('drug_name'),
                dose=con_data.get('dose'),
                route=con_data.get('route'),
                frequency=con_data.get('frequency'),
                therapy_start_date=con_data.get('therapy_start_date'),
                therapy_stop_date=con_data.get('therapy_stop_date'),
                indication=con_data.get('indication'),
            )
            report.concomitant_medications.append(con)

        # Auto-assess causality for each medication individually and save per-drug details
        severity_order = [
            "Certain",
            "Probable/Likely",
            "Possible",
            "Unlikely",
            "Conditional/Unclassified",
            "Unassessable/Unclassifiable",
        ]
        naranjo_order = ["Definite", "Probable", "Possible", "Doubtful"]

        best_who_category = "Unassessable/Unclassifiable"
        best_who_rank = len(severity_order)
        
        best_naranjo_score = -10
        best_naranjo_category = "Doubtful"
        best_naranjo_rank = len(naranjo_order)

        report_dict = data.copy()

        for idx, med in enumerate(report.medications):
            if not med.drug_name:
                continue
            
            # 1. WHO-UMC causality
            who_res = run_assessment(report_dict, med.drug_name)
            who_cat = who_res.get('category', 'Unassessable/Unclassifiable')
            
            # 2. Naranjo causality
            if idx == 0 and data.get('naranjo_answers'):
                naranjo_answers = data.get('naranjo_answers')
            else:
                naranjo_answers = extract_naranjo_features_from_report(report_dict, med.drug_name)
                
            naranjo_res = calculate_naranjo_score(naranjo_answers)
            n_cat = naranjo_res.get('category', 'Doubtful')
            n_score = naranjo_res.get('score', 0)
            
            # Save on SuspectedMedication model
            med.causality_assessment = f"WHO-UMC: {who_cat} | Naranjo: {n_cat} (Score: {n_score})"
            
            # Track best/worst-case WHO-UMC
            who_rank = severity_order.index(who_cat) if who_cat in severity_order else len(severity_order)
            if who_rank < best_who_rank:
                best_who_rank = who_rank
                best_who_category = who_cat
                
            # Track best/worst-case Naranjo
            n_rank = naranjo_order.index(n_cat) if n_cat in naranjo_order else len(naranjo_order)
            if n_rank < best_naranjo_rank:
                best_naranjo_rank = n_rank
                best_naranjo_category = n_cat
            if n_score > best_naranjo_score:
                best_naranjo_score = n_score

        # Save the worst-case summary on the report level
        report.who_umc_category = best_who_category
        report.naranjo_score = best_naranjo_score if best_naranjo_score != -10 else None
        report.naranjo_category = best_naranjo_category

        db.session.add(report)
        db.session.commit()

        return jsonify({
            'success': True,
            'report_id': report.id,
            'who_umc_category': report.who_umc_category,
            'naranjo_score': report.naranjo_score,
            'naranjo_category': report.naranjo_category,
        })

    except Exception as e:
        db.session.rollback()
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 400


@app.route('/api/reports', methods=['GET'])
def get_reports():
    """Get all reports as JSON."""
    reports = ADRReport.query.order_by(ADRReport.created_at.desc()).all()
    return jsonify([r.to_dict() for r in reports])


@app.route('/api/report/<int:report_id>', methods=['GET'])
def get_report(report_id):
    """Get a single report."""
    report = ADRReport.query.get_or_404(report_id)
    return jsonify(report.to_dict())


@app.route('/api/report/<int:report_id>', methods=['DELETE'])
def delete_report(report_id):
    """Delete a report."""
    report = ADRReport.query.get_or_404(report_id)
    db.session.delete(report)
    db.session.commit()
    return jsonify({'success': True})


@app.route('/api/naranjo/calculate', methods=['POST'])
def calculate_naranjo():
    """Calculate Naranjo score from answers."""
    data = request.get_json()
    answers = data.get('answers', [])
    try:
        result = calculate_naranjo_score(answers)
        return jsonify({'success': True, **result})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 400


@app.route('/api/who-umc/assess', methods=['POST'])
def assess_who():
    """Assess causality using WHO-UMC scale."""
    data = request.get_json()
    try:
        result = assess_who_umc(data)
        return jsonify({'success': True, **result})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 400


@app.route('/api/ocr', methods=['POST'])
def ocr_form():
    """
    OCR endpoint: receives a base64 image from the phone camera,
    sends it directly to Gemini using multimodal generation to extract
    structured ADR fields, and returns JSON.
    """
    data = request.get_json()
    image_data = data.get('image')  # base64 string sent from phone camera

    if not image_data:
        return jsonify({'error': 'No image data provided'}), 400

    # Remove data URL prefix if present (e.g. "data:image/jpeg;base64,...")
    if ',' in image_data:
        image_data = image_data.split(',', 1)[1]

    try:
        image_bytes = base64.b64decode(image_data)
    except Exception:
        return jsonify({'error': 'Invalid base64 image data'}), 400

    if not GEMINI_AVAILABLE:
        return jsonify({'error': 'Gemini API is not configured. Set GEMINI_API_KEY environment variable.'}), 503

    try:
        # Step 1 & 2: Use Gemini Multimodal to extract fields directly from the image bytes
        extracted = parse_adr_fields_from_image_with_gemini(image_bytes)
        return jsonify(extracted)

    except Exception as e:
        traceback.print_exc()
        return jsonify({'error': f'OCR processing failed: {str(e)}'}), 500


def parse_adr_fields_from_image_with_gemini(image_bytes):
    """
    Use Gemini Multimodal to parse a handwritten/scanned IPC ADR form image
    directly into structured JSON fields matching the ADR report model.
    """
    prompt = """
You are analyzing an image of an IPC ADR (Adverse Drug Reaction) reporting form.
Read the form contents carefully, perform OCR, extract ALL fields, and return ONLY a valid JSON object. No explanation, no markdown, no extra text.
If a field is not found, use null. Read handwritten text carefully.

IMPORTANT: The form may contain MULTIPLE suspected medications (up to 4 rows in Section C).
Extract ALL of them as an array in the "medications" field.

Fields to extract:

TOP SECTION:
- case_type (string — "initial" if Initial Case is ticked, "follow_up" if Follow-up Case is ticked)
- reg_no (string or null — Reg. No. / IPD / OPD / CR No. from the AMC/NCC section)
- amc_report_no (string or null — AMC Report No.)
- worldwide_unique_no (string or null — Worldwide Unique No.)

A. PATIENT INFORMATION:
- patient_initials (string — field 1)
- patient_age (string, e.g. "35 Years" — field 2)
- gender (string: "male", "female", or "other" — field 3)
- weight_kg (number or null — field 4)

B. SUSPECTED ADVERSE REACTION:
- reaction_start_date (string — field 5, in DD/MM/YYYY format)
- reaction_stop_date (string — field 6, in DD/MM/YYYY format)
- reaction_description (string — field 7, transcribe the FULL handwritten text exactly as written)
- relevant_investigations (string or null — field 12)
- medical_history (string or null — field 13, transcribe full handwritten text)
- is_serious (boolean — field 14, true if "Yes" is ticked)
- seriousness_death (boolean — true if "Death" is ticked in field 14)
- seriousness_death_date (string or null — death date if mentioned, DD/MM/YYYY)
- seriousness_life_threatening (boolean — true if "Life threatening" is ticked)
- seriousness_hospitalization (boolean — true if "Hospitalization/Prolonged" is ticked)
- seriousness_disability (boolean — true if "Disability" is ticked)
- seriousness_congenital_anomaly (boolean — true if "Congenital anomaly" is ticked)
- seriousness_other (boolean — true if "Other Medically important" is ticked)
- outcome (string — field 15: one of "recovered", "recovering", "not_recovered", "fatal", "recovered_with_sequelae", "unknown")

C. SUSPECTED MEDICATION(S):
- medications (array of objects — extract ALL filled rows from field 8 table. Each object must have):
    - drug_name (string — brand or generic name, column 8)
    - manufacturer (string or null — "Manufacturer" column)
    - batch_no (string or null — "Batch No. / Lot No." column)
    - expiry_date (string or null — "Expiry Date" column)
    - dose (string — e.g. "200mg", "40mg", "5mg")
    - route (string — e.g. "ORAL", "IV", "IM")
    - frequency (string — e.g. "OD", "BD", "TDS", "Once daily")
    - therapy_start_date (string or null — "Date Started" column, DD/MM/YYYY)
    - therapy_stop_date (string or null — "Date Stopped" column, DD/MM/YYYY)
    - indication (string or null — "Indication" column, e.g. "POST OP", "PROPHYLAXIS")
    - action_taken (string or null — from field 9 table for this row: one of "drug_withdrawn", "dose_increased", "dose_reduced", "dose_not_changed", "not_applicable", "unknown")
    - reintroduction_result (string or null — from field 10 table for this row: "yes" if reaction reappeared, "no" if no recurrence, "effect_unknown" if effect unknown)
    - reintroduction_dose (string or null — dose if re-introduced, from field 10)

FIELD 11 - CONCOMITANT MEDICATIONS:
- concomitant_medications (array of objects or null — from field 11 table, each with):
    - drug_name (string)
    - dose (string or null)
    - route (string or null)
    - frequency (string or null)
    - therapy_start_date (string or null)
    - therapy_stop_date (string or null)
    - indication (string or null)

ADDITIONAL INFO & SIGNATURES:
- additional_info (string or null — "Additional Information" section)
- receiving_personnel (string or null — "Signature and Name of Receiving Personnel")

D. REPORTER DETAILS:
- reporter_name (string or null — name from field 16)
- reporter_address (string or null — address from field 16, including institution/hospital name)
- reporter_pin (string or null — Pin Code)
- reporter_designation (string or null — Occupation/Designation)
- reporter_email (string or null — Email ID)
- reporter_contact (string or null — Contact No.)
- report_date (string or null — field 17, DD/MM/YYYY)
"""
    image_part = types.Part.from_bytes(
        data=image_bytes,
        mime_type='image/jpeg'
    )
    models_to_try = ['gemini-2.5-flash', 'gemini-2.5-flash-lite', 'gemini-flash-lite-latest']
    response = None
    last_error = None

    for model_name in models_to_try:
        try:
            print(f"DEBUG: Attempting OCR extraction with model {model_name}...")
            response = gemini_client.models.generate_content(
                model=model_name,
                contents=[image_part, prompt]
            )
            print(f"DEBUG: OCR extraction succeeded with model {model_name}.")
            break
        except Exception as e:
            last_error = e
            print(f"WARNING: Model {model_name} failed: {e}. Trying fallback model...")

    if response is None:
        raise last_error

    text = response.text.strip()
    # Strip markdown code fences if present
    if text.startswith("```"):
        text = text.split("```")[1]
        if text.startswith("json"):
            text = text[4:]
    return json.loads(text.strip())


@app.route('/api/analytics/stats', methods=['GET'])
def get_analytics_stats():
    """Get analytics statistics."""
    try:
        total = ADRReport.query.count()
        serious = ADRReport.query.filter_by(is_serious=True).count()

        # Gender distribution
        male = ADRReport.query.filter_by(gender='M').count()
        female = ADRReport.query.filter_by(gender='F').count()
        other_gender = ADRReport.query.filter(ADRReport.gender.notin_(['M', 'F'])).count()

        # Outcome distribution
        outcomes = db.session.query(
            ADRReport.outcome, db.func.count(ADRReport.id)
        ).group_by(ADRReport.outcome).all()

        # WHO-UMC distribution
        who_dist = db.session.query(
            ADRReport.who_umc_category, db.func.count(ADRReport.id)
        ).group_by(ADRReport.who_umc_category).all()

        # Naranjo distribution
        naranjo_dist = db.session.query(
            ADRReport.naranjo_category, db.func.count(ADRReport.id)
        ).group_by(ADRReport.naranjo_category).all()

        # Recent reports by month — handle both SQLite and PostgreSQL
        db_url = app.config.get('SQLALCHEMY_DATABASE_URI', '')
        if 'sqlite' in db_url:
            month_col = db.func.strftime('%Y-%m', ADRReport.created_at)
        else:
            month_col = db.func.to_char(ADRReport.created_at, 'YYYY-MM')

        reports_by_month = db.session.query(
            month_col,
            db.func.count(ADRReport.id)
        ).group_by(month_col).order_by(month_col).limit(12).all()

        # Top drugs
        top_drugs = db.session.query(
            SuspectedMedication.drug_name, db.func.count(SuspectedMedication.id)
        ).filter(SuspectedMedication.drug_name.isnot(None)).group_by(
            SuspectedMedication.drug_name
        ).order_by(db.func.count(SuspectedMedication.id).desc()).limit(10).all()

        return jsonify({
            'total_reports': total,
            'serious_reports': serious,
            'gender_distribution': {
                'male': male, 'female': female, 'other': other_gender
            },
            'outcome_distribution': {str(k or 'Unknown'): v for k, v in outcomes},
            'who_umc_distribution': {str(k or 'N/A'): v for k, v in who_dist},
            'naranjo_distribution': {str(k or 'N/A'): v for k, v in naranjo_dist},
            'reports_by_month': {str(k): v for k, v in reports_by_month},
            'top_drugs': {str(k): v for k, v in top_drugs},
        })
    except Exception as e:
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500


import re

def get_age_group(age_str):
    if not age_str:
        return 'Unknown'
    # Try to find a number in the age string
    match = re.search(r'\d+', str(age_str))
    if not match:
        return 'Unknown'
    try:
        age_val = int(match.group())
        if age_val <= 12:
            return 'Pediatric (0-12)'
        elif age_val <= 19:
            return 'Adolescent (13-19)'
        elif age_val <= 59:
            return 'Adult (20-59)'
        else:
            return 'Geriatric (60+)'
    except:
        return 'Unknown'

@app.route('/api/analytics/dose-age')
def api_dose_age():
    try:
        results = db.session.query(
            SuspectedMedication.drug_name,
            SuspectedMedication.dose,
            ADRReport.patient_age
        ).join(ADRReport, SuspectedMedication.report_id == ADRReport.id).all()
        
        data = {}
        for drug, dose, age_str in results:
            if not drug:
                continue
            drug_clean = drug.strip().upper()
            dose_clean = (dose or 'Unknown').strip()
            
            if drug_clean not in data:
                data[drug_clean] = {}
            if dose_clean not in data[drug_clean]:
                data[drug_clean][dose_clean] = {
                    'count': 0,
                    'age_groups': {
                        'Pediatric (0-12)': 0,
                        'Adolescent (13-19)': 0,
                        'Adult (20-59)': 0,
                        'Geriatric (60+)': 0,
                        'Unknown': 0
                    }
                }
            
            group = get_age_group(age_str)
            data[drug_clean][dose_clean]['count'] += 1
            data[drug_clean][dose_clean]['age_groups'][group] += 1
            
        return jsonify(data)
    except Exception as e:
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500

@app.route('/api/analytics/export-excel')
def api_export_excel():
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from io import BytesIO
        
        wb = Workbook()
        ws_reports = wb.active
        ws_reports.title = "ADR Reports"
        
        # Headers for reports
        headers_reports = [
            "Report ID", "Created At", "Case Type", "Patient Initials", "Age", "Gender", "Weight (kg)",
            "Reaction Start", "Reaction Stop", "Reaction Description", "Is Serious",
            "Seriousness Reasons", "Outcome", "Relevant Investigations", "Medical History",
            "Naranjo Score", "Naranjo Category", "WHO-UMC Category",
            "Reporter Name", "Reporter Occupation", "Reporter Contact"
        ]
        
        ws_reports.append(headers_reports)
        
        # Style headers
        header_font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="8B1A1A", end_color="8B1A1A", fill_type="solid")
        header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
        thin_border = Border(
            left=Side(style='thin', color='DDDDDD'),
            right=Side(style='thin', color='DDDDDD'),
            top=Side(style='thin', color='DDDDDD'),
            bottom=Side(style='thin', color='DDDDDD')
        )
        
        for col_idx in range(1, len(headers_reports) + 1):
            cell = ws_reports.cell(row=1, column=col_idx)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            cell.border = thin_border
            
        reports = ADRReport.query.order_by(ADRReport.id.desc()).all()
        for r in reports:
            ser_list = []
            if r.seriousness_death: ser_list.append("Death")
            if r.seriousness_life_threatening: ser_list.append("Life Threatening")
            if r.seriousness_hospitalization: ser_list.append("Hospitalization")
            if r.seriousness_disability: ser_list.append("Disability")
            if r.seriousness_congenital_anomaly: ser_list.append("Congenital Anomaly")
            if r.seriousness_other: ser_list.append("Other Medically Important")
            ser_str = ", ".join(ser_list) if ser_list else ("Yes" if r.is_serious else "No")
            
            created_str = r.created_at.strftime('%Y-%m-%d %H:%M') if r.created_at else ''
            
            row_data = [
                r.id, created_str, r.case_type, r.patient_initials, r.patient_age, r.gender, r.weight_kg,
                r.reaction_start_date, r.reaction_stop_date, r.reaction_description,
                "Yes" if r.is_serious else "No", ser_str, r.outcome, r.relevant_investigations, r.medical_history,
                r.naranjo_score, r.naranjo_category, r.who_umc_category,
                r.reporter_name, r.reporter_occupation, r.reporter_contact
            ]
            ws_reports.append(row_data)
            
        # Sheet 2: Suspected Medications
        ws_meds = wb.create_sheet(title="Suspected Medications")
        headers_meds = [
            "Medication ID", "Report ID", "Patient Initials", "Drug Name", "Manufacturer", "Batch No",
            "Expiry Date", "Dose", "Route", "Frequency", "Therapy Start", "Therapy Stop", "Indication",
            "Action Taken", "Reintroduction Result"
        ]
        ws_meds.append(headers_meds)
        
        for col_idx in range(1, len(headers_meds) + 1):
            cell = ws_meds.cell(row=1, column=col_idx)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            cell.border = thin_border
            
        meds = SuspectedMedication.query.join(ADRReport).order_by(SuspectedMedication.report_id.desc()).all()
        for m in meds:
            row_data = [
                m.id, m.report_id, m.report.patient_initials if m.report else '', m.drug_name, m.manufacturer, m.batch_no,
                m.expiry_date, m.dose, m.route, m.frequency, m.therapy_start_date, m.therapy_stop_date, m.indication,
                m.action_taken, m.reintroduction_result
            ]
            ws_meds.append(row_data)
            
        # Auto-adjust column widths
        for ws in [ws_reports, ws_meds]:
            for col in ws.columns:
                max_len = 0
                col_letter = col[0].column_letter
                for cell in col:
                    val_str = str(cell.value or '')
                    if len(val_str) > max_len:
                        max_len = len(val_str)
                ws.column_dimensions[col_letter].width = min(max(max_len + 3, 10), 40)
                
        # Style all data rows
        data_font = Font(name="Arial", size=10)
        data_align = Alignment(vertical="center")
        for ws in [ws_reports, ws_meds]:
            for r_idx in range(2, ws.max_row + 1):
                for c_idx in range(1, ws.max_column + 1):
                    cell = ws.cell(row=r_idx, column=c_idx)
                    cell.font = data_font
                    cell.alignment = data_align
                    cell.border = thin_border
                    
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        return send_file(
            output,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            as_attachment=True,
            download_name="adr_reports_export.xlsx"
        )
    except Exception as e:
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500


if __name__ == '__main__':
    app.run(debug=True, port=5000)
